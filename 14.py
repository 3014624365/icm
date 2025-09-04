import openpyxl
from collections import defaultdict

# 打开已存在的Excel文件
source_file = 'athletes.xlsx'  # 假设这是包含原始数据的文件
source_wb = openpyxl.load_workbook(source_file)

# 创建一个新的Excel文件
output_file = 'medal_count_by_year.xlsx'
output_wb = openpyxl.Workbook()

# 遍历输入文件的所有工作表
for sheet_name in source_wb.sheetnames:
    source_ws = source_wb[sheet_name]
    output_ws = output_wb.create_sheet(title=sheet_name)  # 创建一个与原工作表同名的新工作表

    # 用于存储统计结果的字典
    # 格式：{ 'Year': { 'Sport Event': 总奖牌数 } }
    medal_count_by_year = defaultdict(lambda: defaultdict(int))

    # 遍历原始数据，统计奖牌数
    for row in source_ws.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题行
        year = row[2]  # 假设年份在第三列
        event = row[4]  # 假设运动项目在第五列（E列）
        medal = row[6]  # 假设奖牌类型在第七列

        # 检查运动项目列是否有值
        if event and medal in ['Gold', 'Silver', 'Bronze']:
            medal_count_by_year[year][event] += 1

    # 将统计结果写入新的Excel文件
    # 写入标题行
    years = sorted(medal_count_by_year.keys())  # 获取所有年份并排序
    events = set()
    for year, events_dict in medal_count_by_year.items():
        events.update(events_dict.keys())  # 收集所有运动项目
    events = sorted(events)  # 对运动项目排序

    # 写入第一行（运动项目）
    output_ws.cell(row=1, column=1, value="年份")
    for col, event in enumerate(events, start=2):
        output_ws.cell(row=1, column=col, value=event)
    output_ws.cell(row=1, column=len(events) + 2, value="总奖牌数")  # 添加总奖牌数列

    # 写入每个年份的奖牌统计
    for row, year in enumerate(years, start=2):
        output_ws.cell(row=row, column=1, value=year)
        total_medals = 0  # 用于统计该年份的总奖牌数
        for col, event in enumerate(events, start=2):
            medal_count = medal_count_by_year[year].get(event, 0)
            output_ws.cell(row=row, column=col, value=medal_count)
            total_medals += medal_count  # 累加该年份的奖牌数
        output_ws.cell(row=row, column=len(events) + 2, value=total_medals)  # 写入该年份的总奖牌数

# 删除默认创建的Sheet
output_wb.remove(output_wb['Sheet'])

# 保存新的Excel文件
output_wb.save(output_file)
print(f"按年份统计的奖牌结果已保存到 {output_file}")