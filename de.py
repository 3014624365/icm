import pandas as pd
from openpyxl import load_workbook

# 需要删除的工作表名称列表
sheets_to_delete = [
    "Artistic Swimming", "Diving", "Marathon Swimming", "Swimming", "Water Polo", "Archery", "Athletics", "Badminton",
    "Baseball", "Softball", "3x3", "Basketball", "Basque Pelota", "Boxing", "Breaking", "Sprint", "Slalom", "Cricket",
    "Croquet", "BMX Freestyle", "BMX Racing", "Mountain Bike", "Road", "Track", "Dressage", "Eventing", "Jumping",
    "Vaulting", "Driving", "Fencing", "Field hockey", "Flag football", "Football", "Golf", "Artistic", "Rhythmic",
    "Trampoline", "Indoor", "Field", "Jeu de Paume", "Judo", "Karate", "Sixes", "Field.1", "Unnamed: 45", "Polo",
    "Rackets", "Roque", "Coastal", "Rowing", "Sevens", "Union", "Sailing", "Shooting", "Skateboarding", "Sport Climbing",
    "Squash", "Surfing", "Table Tennis", "Taekwondo", "Tennis", "Triathlon", "Tug of War", "Beach", "Indoor.1",
    "Unnamed: 66", "Weightlifting", "Freestyle", "Greco-Roman", "Figure", "Ice Hockey"
]

# 加载 Excel 文件
file_path = 'country_awards_ratio.xlsx'
wb = load_workbook(file_path)

# 遍历所有工作表，删除指定名称的工作表
for sheet_name in sheets_to_delete:
    if sheet_name in wb.sheetnames:  # 检查工作表是否存在
        wb.remove(wb[sheet_name])  # 删除工作表
        print(f"已删除工作表: {sheet_name}")
    else:
        print(f"工作表不存在: {sheet_name}")

# 保存修改后的 Excel 文件
wb.save(file_path)
print(f"文件已保存: {file_path}")